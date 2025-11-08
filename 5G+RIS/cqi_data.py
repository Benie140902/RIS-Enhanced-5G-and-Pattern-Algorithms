import socket, json
from datetime import datetime
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from collections import deque
from matplotlib.animation import FuncAnimation
import numpy as np

UDP_IP = "127.0.0.1"
UDP_PORT = 55555
OUT_XLSX = "cqi_log_pos.xlsx"
BUFFER = 65536

# Store last N points for plotting
MAX_POINTS = 200
timestamps = deque(maxlen=MAX_POINTS)
cqis = deque(maxlen=MAX_POINTS)
data_rates = deque(maxlen=MAX_POINTS)

# Track last CQI
last_cqi = [None]

# Helper: recursively search JSON for (rnti -> CQI) entries
def find_cqi_entries(obj, path=""):
    entries = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if isinstance(v, dict) and any(key.lower() == "cqi" for key in v.keys()):
                cqi_val = None
                for kk in v.keys():
                    if kk.lower() == "cqi":
                        cqi_val = v[kk]
                entries.append((k, cqi_val, f"{path}/{k}"))
            entries += find_cqi_entries(v, path + "/" + str(k))
    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            entries += find_cqi_entries(item, path + f"/[{i}]")
    return entries

# Helper: recursively find throughput keys
def find_datarate_entries(obj, path=""):
    rates = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if isinstance(v, (int, float)) and ("brate" in k.lower() or "throughput" in k.lower()):
                rates.append((k, v))
            rates += find_datarate_entries(v, path + "/" + str(k))
    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            rates += find_datarate_entries(item, path + f"/[{i}]")
    return rates

# Load or create workbook
#try:
 #   wb = load_workbook(OUT_XLSX)
  #  ws = wb.active
#except FileNotFoundError:
 #   wb = Workbook()
  #  ws = wb.active
   # ws.append(["timestamp_iso", "rnti", "cqi", "raw_json_fragment", "data_rate_kbps"])
    #wb.save(OUT_XLSX)

# Open UDP socket
sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
sock.bind((UDP_IP, UDP_PORT))
sock.setblocking(False)

#print(f"Listening for JSON metrics on {UDP_IP}:{UDP_PORT} -> logging to {OUT_XLSX}")

# Matplotlib setup
plt.style.use("seaborn-v0_8-darkgrid")


fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 8), sharex=True)

# CQI plot (continuous blue)
line_cqi, = ax1.plot([], [], color="blue", linewidth=2, label="CQI")
ax1.set_ylim(0, 16)
ax1.set_ylabel("CQI")
ax1.legend(loc="upper right")

# Data rate plot (in Mbps)
line_rate, = ax2.plot([], [], color="green", linewidth=2, label="Data Rate (Mbps)")
ax2.set_ylabel("Data Rate (Kbps)")
ax2.legend(loc="upper right")

# Text annotations for live values
text_cqi = ax1.text(0.98, 0.9, "", transform=ax1.transAxes,
                    ha="right", va="top", fontsize=10, color="blue")
text_rate = ax2.text(0.98, 0.9, "", transform=ax2.transAxes,
                     ha="right", va="top", fontsize=10, color="green")

def animate(frame):
    try:
        data, addr = sock.recvfrom(BUFFER)
        now = datetime.utcnow().isoformat() + "Z"
        j = json.loads(data.decode("utf-8"))

        found = []
        if isinstance(j, dict) and "ue_data" in j and isinstance(j["ue_data"], dict):
            for rnti, ue_obj in j["ue_data"].items():
                cqi_val = None
                for kk, vv in ue_obj.items():
                    if kk.lower() == "cqi":
                        cqi_val = vv
                        break
                found.append((rnti, cqi_val, f"ue_data/{rnti}"))

        if not found:
            found = find_cqi_entries(j)

        # Extract datarate
        rates_found = find_datarate_entries(j)
        rate_val_kbps = None
        if rates_found:
            rate_val_kbps = rates_found[0][1]  # first match

        for rnti, cqi_val, p in found:
            if cqi_val is None:
                continue

            timestamps.append(now)
            cqis.append(cqi_val)
            data_rates.append(rate_val_kbps / 1000.0 if rate_val_kbps is not None else np.nan)  # store in Mbps
            last_cqi[0] = cqi_val

           # ws.append([now, rnti, cqi_val, json.dumps({p: cqi_val}), rate_val_kbps])
            #wb.save(OUT_XLSX)

    except BlockingIOError:
        pass  # No packet received this frame

    # Prepare X and Y values
    xs = list(range(len(cqis)))
    cqis_list = list(cqis)
    rate_list = list(data_rates)

    line_cqi.set_data(xs, cqis_list)
    line_rate.set_data(xs, rate_list)

    ax1.set_xlim(max(0, len(cqis) - MAX_POINTS), len(cqis))
    ax2.set_xlim(max(0, len(cqis) - MAX_POINTS), len(cqis))

    if rate_list:
        ax2.set_ylim(0, max(1, np.nanmax(rate_list) * 1.2))  # scale Mbps

    # Update text annotations with latest values
    if cqis_list:
        text_cqi.set_text(f"CQI: {cqis_list[-1]:.0f}")
    if rate_list and not np.isnan(rate_list[-1]):
        text_rate.set_text(f"Rate: {rate_list[-1]:.2f} Kbps")

    return line_cqi, line_rate, text_cqi, text_rate

ani = FuncAnimation(fig, animate, interval=500)
plt.show()
